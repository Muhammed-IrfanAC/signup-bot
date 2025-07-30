# Event-related API routes.
from flask import Blueprint, request, jsonify, send_file
import io
from firebase_admin import firestore
import requests
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ... import Config

# Create blueprint
events_bp = Blueprint('events', __name__)
db = firestore.client()

def log_event_action(guild_id: str, event_name: str, action: str, user_name: str, 
                    user_avatar_url: str, success: bool, details: str = "", 
                    error_reason: str = "", additional_data: dict = None):
    """Log an event action to the designated log channel."""
    try:
        # Get the log channel ID from the event data
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        event_doc = event_ref.get()
        
        if not event_doc.exists:
            return
            
        event_data = event_doc.to_dict()
        log_channel_id = event_data.get('log_channel_id')
        
        if not log_channel_id:
            return
            
        # Store log entry in Firebase for the bot to process
        log_entry = {
            'guild_id': guild_id,
            'event_name': event_name,
            'action': action,
            'user_name': user_name,
            'user_avatar_url': user_avatar_url,
            'success': success,
            'details': details,
            'error_reason': error_reason,
            'additional_data': additional_data or {},
            'timestamp': datetime.utcnow().isoformat(),
            'processed': False
        }
        
        # Store in a logs collection
        db.collection('servers').document(str(guild_id)).collection('events').document(event_name).collection('logs').add(log_entry)
        
    except Exception as e:
        print(f"Error logging action: {e}")

def is_user_leader(guild_id: str, user_roles: list) -> bool:
    """Check if user has any leader role."""
    try:
        # Get all leader roles for the guild
        leader_roles_ref = db.collection('servers').document(str(guild_id)).collection('server_leaders').document('roles')
        leader_roles_doc = leader_roles_ref.get()
        
        if not leader_roles_doc.exists:
            return False
            
        leader_roles = leader_roles_doc.to_dict().get('leader_role_ids', [])
        
        # Check if any of user's roles is in leader_roles
        return any(role_id in leader_roles for role_id in user_roles)
        
    except Exception as e:
        print(f"Error checking leader role: {e}")
        return False

def player_get(player_tag: str) -> dict:
    try:
        url = f'https://cocproxy.royaleapi.dev/v1/players/%23{player_tag.lstrip("#")}'
        response = requests.get(url, headers={'Authorization': Config.AUTH})
        if response.status_code == 200:
            return response.json()
        else:
            return {} 
    except Exception as e:
        print(f"Error fetching player data: {e}")
        return {}

@events_bp.route('', methods=['POST'])
def create_event():
    """Create a new event."""
    try:
        data = request.json
        event_name = data.get('event_name')
        guild_id = data.get('guild_id')
        channel_id = data.get('channel_id')
        role_id = data.get('role_id')  # Optional role ID
        log_channel_id = data.get('log_channel_id')  # Optional log channel ID
        
        if not event_name or not guild_id:
            return jsonify({'error': 'Event name and guild ID are required'}), 400
        
        # Check if event already exists
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        if event_ref.get().exists:
            return jsonify({'error': 'Event with this name already exists'}), 400
        
        # Create the event
        event_data = {
            'event_name': event_name,
            'created_at': datetime.utcnow().isoformat(),
            'signup_count': 0,
            'is_open': True,
            'embed': {
                'title': event_name,
                'description': 'Event roster and signups',
                'fields': [
                    {'name': 'Total Signups', 'value': '0', 'inline': False},
                    {'name': 'TH Composition', 'value': 'No signups yet', 'inline': False}
                ],
                'color': 0x00ff00
            },
            'channel_id': channel_id
        }
        
        # Add role_id if provided
        if role_id:
            event_data['role_id'] = role_id
            
        # Add log_channel_id if provided
        if log_channel_id:
            event_data['log_channel_id'] = log_channel_id
        
        event_ref.set(event_data)
        
        # Log the event creation
        user_name = data.get('user_name', 'Unknown User')
        user_avatar_url = data.get('user_avatar_url', '')
        log_event_action(
            guild_id=str(guild_id),
            event_name=event_name,
            action='create',
            user_name=user_name,
            user_avatar_url=user_avatar_url,
            success=True,
            details=f"Event '{event_name}' created successfully",
            additional_data={
                'channel_id': channel_id,
                'role_id': role_id,
                'log_channel_id': log_channel_id
            }
        )
        
        return jsonify({'message': 'Event created successfully', 'event_name': event_name}), 201
        
    except Exception as e:
        # Log the error
        user_name = request.json.get('user_name', 'Unknown User') if request.json else 'Unknown User'
        user_avatar_url = request.json.get('user_avatar_url', '') if request.json else ''
        log_event_action(
            guild_id=str(request.json.get('guild_id', '')) if request.json else '',
            event_name=request.json.get('event_name', '') if request.json else '',
            action='create',
            user_name=user_name,
            user_avatar_url=user_avatar_url,
            success=False,
            error_reason=str(e)
        )
        return jsonify({'error': str(e)}), 500

@events_bp.route('', methods=['GET'])
def list_events():
    """List all events for a guild."""
    try:
        guild_id = request.args.get('guild_id')
        if not guild_id:
            return jsonify({'error': 'Guild ID is required'}), 400
        
        # Get all events for the guild
        events_ref = db.collection('servers').document(str(guild_id)).collection('events')
        events = []
        
        for doc in events_ref.stream():
            event_data = doc.to_dict()
            event_data['id'] = doc.id
            events.append(event_data)
        
        return jsonify({'events': events}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@events_bp.route('/<event_name>', methods=['GET'])
def get_event(event_name):
    """Get details for a specific event including signups."""
    try:
        guild_id = request.args.get('guild_id')
        if not guild_id:
            return jsonify({'error': 'Guild ID is required'}), 400
        
        # Get event details
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        event_doc = event_ref.get()
        
        if not event_doc.exists:
            return jsonify({'error': 'Event not found'}), 404
            
        event_data = event_doc.to_dict()
        event_data['id'] = event_doc.id
        
        # Get signups for the event
        signups_ref = event_ref.collection('signups').order_by('index')
        signups = [doc.to_dict() for doc in signups_ref.stream()]
        
        # Calculate TH composition
        th_composition = {}
        for signup in signups:
            th = signup.get('player_th', 0)
            th_composition[th] = th_composition.get(th, 0) + 1
        
        # Add signups and composition to response
        event_data['signups'] = signups
        event_data['th_composition'] = th_composition
        event_data['signup_count'] = len(signups)
            
        return jsonify(event_data), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@events_bp.route('/<event_name>/signup', methods=['POST'])
def signup_player(event_name):
    """Sign up a player for an event."""
    try:
        data = request.json
        player_tag = data.get('player_tag')
        discord_name = data.get('discord_name')
        guild_id = data.get('guild_id')
        discord_user_id = data.get('discord_user_id')  # Add Discord user ID
        
        if not all([player_tag, discord_name, guild_id]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Check if event exists and is open
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        event_doc = event_ref.get()
        
        if not event_doc.exists:
            # Log the error
            log_event_action(
                guild_id=str(guild_id),
                event_name=event_name,
                action='signup',
                user_name=discord_name,
                user_avatar_url=data.get('user_avatar_url', ''),
                success=False,
                error_reason='Event not found'
            )
            return jsonify({'error': 'Event not found'}), 404
            
        if not event_doc.to_dict().get('is_open', True):
            # Log the error
            log_event_action(
                guild_id=str(guild_id),
                event_name=event_name,
                action='signup',
                user_name=discord_name,
                user_avatar_url=data.get('user_avatar_url', ''),
                success=False,
                error_reason='Event registration is closed'
            )
            return jsonify({'error': 'Event registration is closed'}), 400
        
        # Check if player is already signed up
        signups_ref = event_ref.collection('signups')
        existing_signup = signups_ref.where('player_tag', '==', player_tag).limit(1).get()
        
        if len(existing_signup) > 0:
            # Log the error
            log_event_action(
                guild_id=str(guild_id),
                event_name=event_name,
                action='signup',
                user_name=discord_name,
                user_avatar_url=data.get('user_avatar_url', ''),
                success=False,
                error_reason='Player already signed up for this event'
            )
            return jsonify({'error': 'You are already signed up for this event'}), 400
        
        # Get player data from Clash of Clans API
        try:
            player_data = player_get(player_tag)
            if player_data:
                player_name = player_data.get('name', 'Unknown')
                player_th = player_data.get('townHallLevel', 0)
            else:
                raise Exception('Please check the player tag.')
        except Exception as e:
            # Log the error
            log_event_action(
                guild_id=str(guild_id),
                event_name=event_name,
                action='signup',
                user_name=discord_name,
                user_avatar_url=data.get('user_avatar_url', ''),
                success=False,
                error_reason=f'Failed to fetch player data: {str(e)}'
            )
            return jsonify({'error': 'Failed to fetch player data. Please check the player tag.'}), 400
        
        # Add signup
        signup_data = {
            'player_name': player_name,
            'player_tag': player_tag,
            'player_th': player_th,
            'discord_name': discord_name,
            'discord_user_id': discord_user_id,  # Store Discord user ID
            'signed_up_at': datetime.utcnow().isoformat()
        }
        
        # Get next index
        signup_count = event_doc.to_dict().get('signup_count', 0)
        signup_data['index'] = signup_count + 1
        
        # Add to database
        signups_ref.add(signup_data)
        event_ref.update({'signup_count': signup_count + 1})
        
        # Get event data to check for role_id
        event_data = event_doc.to_dict()
        role_id = event_data.get('role_id')
        
        # Log successful signup
        log_event_action(
            guild_id=str(guild_id),
            event_name=event_name,
            action='signup',
            user_name=discord_name,
            user_avatar_url=data.get('user_avatar_url', ''),
            success=True,
            details=f"Player {player_name} (TH{player_th}) signed up successfully",
            additional_data={
                'player_name': player_name,
                'player_tag': player_tag,
                'player_th': player_th,
                'signup_index': signup_count + 1
            }
        )
        
        return jsonify({
            'message': 'Signup successful',
            'player_name': player_name,
            'player_th': player_th,
            'role_id': role_id,  # Return role_id if it exists
            'discord_user_id': discord_user_id
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@events_bp.route('/<event_name>/signups', methods=['GET'])
def get_signups(event_name):
    """Get all signups for an event."""
    try:
        guild_id = request.args.get('guild_id')
        if not guild_id:
            return jsonify({'error': 'Guild ID is required'}), 400
        
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        if not event_ref.get().exists:
            return jsonify({'error': 'Event not found'}), 404
            
        signups = []
        for doc in event_ref.collection('signups').order_by('index').stream():
            signup = doc.to_dict()
            signup['id'] = doc.id
            signups.append(signup)
            
        return jsonify({'signups': signups, 'count': len(signups)}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@events_bp.route('/<event_name>/export', methods=['GET'])
def export_event(event_name):
    """Export event data to Excel."""
    try:
        guild_id = request.args.get('guild_id')
        if not guild_id:
            return jsonify({'error': 'Guild ID is required'}), 400
        
        # Get event data
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        event_doc = event_ref.get()
        
        if not event_doc.exists:
            return jsonify({'error': 'Event not found'}), 404
        
        # Get signups
        signups = []
        for doc in event_ref.collection('signups').order_by('index').stream():
            signups.append(doc.to_dict())
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = event_name[:31]  # Excel sheet name limit
        
        # Add headers
        headers = ['#', 'Player Name', 'Player Tag', 'TH Level', 'Discord Name', 'Signed Up At']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = Font(bold=True)
        
        # Add data
        for row_num, signup in enumerate(signups, 2):
            ws[f'A{row_num}'] = signup.get('index', '')
            ws[f'B{row_num}'] = signup.get('player_name', '')
            ws[f'C{row_num}'] = signup.get('player_tag', '')
            ws[f'D{row_num}'] = signup.get('player_th', '')
            ws[f'E{row_num}'] = signup.get('discord_name', '')
            ws[f'F{row_num}'] = signup.get('signed_up_at', '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        # Log successful export
        log_event_action(
            guild_id=str(guild_id),
            event_name=event_name,
            action='export',
            user_name=request.args.get('user_name', 'Unknown User'),
            user_avatar_url=request.args.get('user_avatar_url', ''),
            success=True,
            details=f"Event '{event_name}' data exported successfully",
            additional_data={
                'signup_count': len(signups),
                'file_name': f"{event_name}_export.xlsx"
            }
        )
        
        return send_file(
            file_stream,
            as_attachment=True,
            download_name=f"{event_name}_export.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@events_bp.route('/<event_name>/close', methods=['POST'])
async def close_event(event_name):
    """Close event registration."""
    try:
        guild_id = request.json.get('guild_id')
        if not guild_id:
            return jsonify({'error': 'Guild ID is required'}), 400
        
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        if not event_ref.get().exists:
            return jsonify({'error': 'Event not found'}), 404

        if not is_user_leader(guild_id, request.json.get('user_roles', [])):
            # Log the error
            log_event_action(
                guild_id=str(guild_id),
                event_name=event_name,
                action='close',
                user_name=request.json.get('user_name', 'Unknown User'),
                user_avatar_url=request.json.get('user_avatar_url', ''),
                success=False,
                error_reason='User does not have leader permissions'
            )
            return jsonify({'error': 'You must be a leader to close an event'}), 403
            
        event_ref.update({'is_open': False})
        
        # Log successful closure
        log_event_action(
            guild_id=str(guild_id),
            event_name=event_name,
            action='close',
            user_name=request.json.get('user_name', 'Unknown User'),
            user_avatar_url=request.json.get('user_avatar_url', ''),
            success=True,
            details=f"Event '{event_name}' registration closed successfully"
        )
        
        return jsonify({'message': 'Event registration closed successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@events_bp.route('/<event_name>/check', methods=['POST'])
def check_player(event_name):
    """Check if a player is signed up for an event."""
    try:
        data = request.json
        player_tag = data.get('player_tag')
        guild_id = data.get('guild_id')
        
        if not player_tag or not guild_id:
            return jsonify({'error': 'Player tag and guild ID are required'}), 400
        
        # Check if event exists
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        if not event_ref.get().exists:
            return jsonify({'error': 'Event not found'}), 404
        
        # Check if player is signed up
        signups_ref = event_ref.collection('signups')
        signup_docs = signups_ref.where('player_tag', '==', player_tag).limit(1).stream()
        signup_docs = list(signup_docs)
        
        if not signup_docs:
            return jsonify({
                'is_signed_up': False,
                'message': 'Player is not signed up for this event'
            }), 200
        
        # Get player details if signed up
        signup_data = signup_docs[0].to_dict()
        return jsonify({
            'is_signed_up': True,
            'message': 'Player is signed up for this event',
            'player_data': {
                'name': signup_data.get('player_name'),
                'th_level': signup_data.get('player_th'),
                'discord_name': signup_data.get('discord_name'),
                'index': signup_data.get('index')
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@events_bp.route('/<event_name>/remove', methods=['POST'])
async def remove_player(event_name):
    """Remove a player from an event."""
    try:
        data = request.json
        player_tag = data.get('player_tag')
        discord_name = data.get('discord_name')
        guild_id = data.get('guild_id')
        
        if not guild_id:
            return jsonify({'error': 'Guild ID is required'}), 400
            
        # Check if user is a leader
        is_leader = is_user_leader(guild_id, data.get('user_roles', []))
        
        if not all([player_tag, discord_name, guild_id]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Check if event exists
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        event_doc = event_ref.get()
        
        if not event_doc.exists:
            return jsonify({'error': 'Event not found'}), 404
        
        # Find the player's signup
        signups_ref = event_ref.collection('signups')
        signup_query = signups_ref.where('player_tag', '==', player_tag).limit(1)
        signup_docs = list(signup_query.stream())
        
        if not signup_docs:
            return jsonify({'error': 'Player not found in this event'}), 404
            
        signup_doc = signup_docs[0]
        signup_data = signup_doc.to_dict()
        
        # Check permissions
        if not is_leader and signup_data.get('discord_name') != discord_name:
            return jsonify({'error': 'You can only remove your own signup'}), 403
        
        # Store player data for response
        player_data = {
            'name': signup_data.get('player_name'),
            'th_level': signup_data.get('player_th'),
            'discord_user_id': signup_data.get('discord_user_id')
        }
        
        # Get event data to check for role_id
        event_data = event_doc.to_dict()
        role_id = event_data.get('role_id')
        
        # Get the index of the player being removed
        removed_index = signup_data.get('index', 0)
        
        # Start a batch write
        batch = db.batch()
        
        # Delete the signup
        batch.delete(signup_doc.reference)
        
        # Update indexes of all signups with higher index
        higher_signups = signups_ref.where('index', '>', removed_index).stream()
        for doc in higher_signups:
            doc_ref = doc.reference
            new_data = doc.to_dict()
            new_data['index'] = new_data['index'] - 1
            batch.update(doc_ref, new_data)
        
        # Update total count
        current_count = event_doc.to_dict().get('signup_count', 0)
        batch.update(event_ref, {'signup_count': current_count - 1})
        
        # Commit all updates
        batch.commit()
        
        # Log successful removal
        log_event_action(
            guild_id=str(guild_id),
            event_name=event_name,
            action='remove',
            user_name=discord_name,
            user_avatar_url=data.get('user_avatar_url', ''),
            success=True,
            details=f"Player {player_data['name']} (TH{player_data['th_level']}) removed from event",
            additional_data={
                'player_name': player_data['name'],
                'player_th': player_data['th_level'],
                'removed_index': removed_index,
                'is_self_removal': signup_data.get('discord_name') == discord_name
            }
        )
        
        return jsonify({
            'message': 'Player removed successfully',
            'player_data': player_data,
            'role_id': role_id,  # Return role_id if it exists
            'is_self_removal': signup_data.get('discord_name') == discord_name  # Check if it's self-removal
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@events_bp.route('/<event_name>/update_message_id', methods=['POST'])
def update_message_id(event_name):
    """Update the message ID for an event's embed."""
    try:
        data = request.json
        guild_id = data.get('guild_id')
        message_id = data.get('message_id')
        channel_id = data.get('channel_id')
        
        if not guild_id or not message_id:
            return jsonify({'error': 'Guild ID and message ID are required'}), 400
        
        # Update the message ID (and channel ID if provided) in Firestore
        event_ref = db.collection('servers').document(str(guild_id)).collection('events').document(event_name)
        if not event_ref.get().exists:
            return jsonify({'error': 'Event not found'}), 404
        update_data = {'message_id': message_id}
        if channel_id:
            update_data['channel_id'] = channel_id
        event_ref.update(update_data)
        return jsonify({'message': 'Message ID updated successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
